from fastapi import APIRouter, UploadFile, File, HTTPException
from db import get_pool
from openpyxl import load_workbook
from datetime import date
import io, json

router = APIRouter()

REQUIRED_COLS = {"student_id", "subject", "test_name", "test_date", "marks_obtained", "total_marks"}


def parse_excel(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    records = []
    errors  = []
    for i, row in enumerate(rows[1:], start=2):
        r = dict(zip(headers, row))
        if all(r.get(c) is None for c in REQUIRED_COLS):
            continue  # skip blank rows

        # Validate & coerce
        row_errors = []
        if not r.get("student_id"):
            row_errors.append("student_id missing")
        try:
            marks = float(r["marks_obtained"])
            total = float(r["total_marks"])
            if marks < 0:       row_errors.append("marks_obtained < 0")
            if total <= 0:      row_errors.append("total_marks must be > 0")
            if marks > total:   row_errors.append("marks_obtained > total_marks")
        except (TypeError, ValueError):
            row_errors.append("marks_obtained / total_marks must be numbers")
            marks, total = 0, 1

        # Accept date objects (Excel native) or strings
        td = r.get("test_date")
        if isinstance(td, date):
            test_date = td
        elif isinstance(td, str):
            try:
                test_date = date.fromisoformat(td.strip())
            except ValueError:
                row_errors.append(f"test_date '{td}' not YYYY-MM-DD")
                test_date = None
        else:
            row_errors.append("test_date missing or invalid")
            test_date = None

        if row_errors:
            errors.append({"row": i, "errors": row_errors})
            continue

        records.append({
            "student_id":     str(r["student_id"]).strip(),
            "subject":        str(r["subject"]).strip(),
            "test_name":      str(r["test_name"]).strip(),
            "test_date":      test_date,
            "marks_obtained": marks,
            "total_marks":    total,
        })

    return records, errors


@router.post("/upload/scores/preview")
async def preview(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files accepted")
    data = await file.read()
    records, errors = parse_excel(data)
    # Return preview (no DB write yet)
    return {
        "valid":  len(records),
        "errors": errors,
        "rows": [
            {**r, "test_date": r["test_date"].isoformat()}
            for r in records[:50]   # cap preview at 50 rows
        ],
    }


@router.post("/upload/scores/confirm")
async def confirm(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files accepted")
    data = await file.read()
    records, errors = parse_excel(data)

    if not records:
        raise HTTPException(status_code=400, detail="No valid rows to import")

    pool = await get_pool()
    inserted = skipped = 0

    async with pool.acquire() as conn:
        # Verify student IDs exist
        ids = list({r["student_id"] for r in records})
        existing_ids = {
            row["id"] for row in await conn.fetch(
                "SELECT id FROM students WHERE id = ANY($1) AND is_deleted = FALSE", ids
            )
        }

        async with conn.transaction():
            for r in records:
                if r["student_id"] not in existing_ids:
                    skipped += 1
                    continue
                await conn.execute("""
                    INSERT INTO test_scores
                        (student_id, subject, test_name, test_date, marks_obtained, total_marks, imported_from)
                    VALUES ($1, $2, $3, $4, $5, $6, 'excel')
                    ON CONFLICT (student_id, subject, test_name, test_date)
                    DO UPDATE SET
                        marks_obtained  = EXCLUDED.marks_obtained,
                        total_marks     = EXCLUDED.total_marks,
                        previous_marks  = test_scores.marks_obtained,
                        imported_from   = 'excel'
                """,
                    r["student_id"], r["subject"], r["test_name"],
                    r["test_date"], r["marks_obtained"], r["total_marks"]
                )

                await conn.execute("""
                    INSERT INTO change_log
                        (table_name, record_id, action, new_values, changed_by, note)
                    VALUES ('test_scores', $1, 'insert', $2::jsonb, 'excel_import', 'Imported via dashboard upload')
                """,
                    r["student_id"],
                    json.dumps({**r, "test_date": r["test_date"].isoformat()})
                )
                inserted += 1

    return {"inserted": inserted, "skipped_unknown_ids": skipped, "parse_errors": len(errors)}
